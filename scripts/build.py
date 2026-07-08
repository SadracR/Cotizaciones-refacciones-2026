"""
build.py — Genera index.html a partir del archivo Excel de control.
Se ejecuta automáticamente por GitHub Actions cada vez que
se sube un archivo .xlsx al repositorio.
"""

import openpyxl
import json
import datetime
import glob
import sys
import os

# ── Buscar el archivo Excel ──────────────────────────────────────────────────
EXCEL_CANDIDATES = glob.glob("*.xlsx") + glob.glob("CONTROL*.xlsx")
if not EXCEL_CANDIDATES:
    print("ERROR: No se encontró ningún archivo .xlsx en la raíz del repositorio.")
    sys.exit(1)

EXCEL_FILE = EXCEL_CANDIDATES[0]
print(f"Leyendo: {EXCEL_FILE}")

# ── Utilidades ────────────────────────────────────────────────────────────────
def clean(v):
    if v is None: return None
    if isinstance(v, str): t = v.strip(); return t or None
    return v

def num(v):
    return v if isinstance(v, (int, float)) and v > 0 else None

def fmt_date(v):
    if isinstance(v, datetime.datetime): return v.strftime("%Y-%m-%d")
    return None

# ── Leer Excel ────────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)

# Hoja CONTROL
if "CONTROL" not in wb.sheetnames:
    print(f"ERROR: No se encontró la hoja 'CONTROL' en {EXCEL_FILE}")
    print(f"Hojas disponibles: {wb.sheetnames}")
    sys.exit(1)

ws = wb["CONTROL"]

# Hoja MAQUINARIA (opcional)
maq = {}
if "MAQUINARIA" in wb.sheetnames:
    wm = wb["MAQUINARIA"]
    for row in wm.iter_rows(min_row=2, max_row=700, min_col=1, max_col=14, values_only=True):
        eco = row[1]
        if not eco or not isinstance(row[0], (int, float)): continue
        if abs(row[0] - round(row[0])) > 0.01: continue
        maq[str(eco)] = {
            "familia":  clean(row[2]) or "",
            "marca":    clean(row[3]) or "",
            "modelo":   str(row[4]).strip() if row[4] else "",
            "obra":     clean(row[5]) or "",
            "status":   str(row[7] or "").strip().upper(),
        }
    print(f"Maquinaria: {len(maq)} equipos")

# Filas de CONTROL
rows = []
for row in ws.iter_rows(min_row=2, max_row=2000, min_col=1, max_col=33, values_only=True):
    if all(v is None for v in row): continue
    status = str(row[20] or "").strip().upper()
    tema   = str(row[32] or "").strip()
    es_srv = tema.upper() == "SERVICIO" or (not row[8] and bool(tema))
    rows.append({
        "pedido":     row[0],
        "fecha":      fmt_date(row[1]),
        "solicita":   clean(row[2]),
        "econStatus": str(row[6] or "").strip().upper(),
        "numEcon":    clean(row[7]),
        "parte":      str(row[8]).strip() if row[8] else None,
        "desc":       clean(row[9]),
        "cantidad":   row[10] if isinstance(row[10], (int, float)) else 0,
        "matcoUnit":  num(row[11]),
        "matcoTotal": num(row[12]),
        "matcoDias":  clean(str(row[13])) if row[13] else None,
        "hawkUnit":   num(row[14]),
        "hawkTotal":  num(row[15]),
        "hawkDias":   clean(str(row[16])) if row[16] else None,
        "ctpUnit":    num(row[17]),
        "ctpTotal":   num(row[18]),
        "ctpDias":    clean(str(row[19])) if row[19] else None,
        "status":     status,
        "cotizacion": clean(row[21]),
        "tema":       tema,
        "esServicio": es_srv,
    })

print(f"Filas leídas: {len(rows)}")
cotizando = [r for r in rows if r["status"] == "COTIZANDO"]
print(f"COTIZANDO: {len(cotizando)}")

# ── Generar JSON ──────────────────────────────────────────────────────────────
data_json = json.dumps({"rows": rows, "maq": maq}, ensure_ascii=False)

# ── Leer template ─────────────────────────────────────────────────────────────
template_path = os.path.join(os.path.dirname(__file__), "template.html")
with open(template_path, "r", encoding="utf-8") as f:
    template = f.read()

if "/*__DATA__*/" not in template:
    print("ERROR: El template.html no contiene el marcador /*__DATA__*/")
    sys.exit(1)

# ── Inyectar datos y metadatos ────────────────────────────────────────────────
final = template.replace("/*__DATA__*/", data_json)

# Agregar fecha de última actualización en el footer
fecha_hoy = datetime.datetime.utcnow().strftime("%d/%m/%Y %H:%M UTC")
final = final.replace(
    "Control de Compras — Refacciones 2026 · Solo lectura",
    f"Control de Compras — Refacciones 2026 · Actualizado: {fecha_hoy}"
)

# ── Escribir index.html ───────────────────────────────────────────────────────
with open("index.html", "w", encoding="utf-8") as f:
    f.write(final)

size_kb = round(len(final) / 1024)
print(f"index.html generado correctamente ({size_kb} KB)")
print(f"Fuente: {EXCEL_FILE} → {len(rows)} partidas, {len(cotizando)} en COTIZANDO")
